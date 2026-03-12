"""Standalone script — run from terminal without Excel add-in.

Usage:
    python -m isrc_fetcher.standalone "path/to/file.xlsx" [--rows 2-10] [--all]
"""
from __future__ import annotations

import argparse
import sys
import re

import openpyxl

from isrc_fetcher.columns import (
    COL_DURACION as COL_B,
    COL_TITULO_GRABACION as COL_C,
    COL_INTERPRETES as COL_D,
    COL_ISRC as COL_H,
    COL_STATUS, COL_FOUND_TITLE, COL_FOUND_ARTIST, COL_FOUND_DURATION,
)
from isrc_fetcher.config import load_config, save_config
from isrc_fetcher.fetcher import ISRCFetcher


def parse_duration(raw) -> int | None:
    if raw is None:
        return None
    text = str(raw).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", text)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return None


def main():
    parser = argparse.ArgumentParser(description="Fetch ISRC codes for songs in an Excel file")
    parser.add_argument("file", help="Path to the Excel (.xlsx) file")
    parser.add_argument("--rows", help="Row range to process, e.g. '2-10' or '5' (1-based, header=1)")
    parser.add_argument("--all", action="store_true", help="Process all data rows")
    parser.add_argument("--spotify-id", help="Spotify Client ID (or set in config)")
    parser.add_argument("--spotify-secret", help="Spotify Client Secret (or set in config)")
    parser.add_argument("--save-credentials", action="store_true", help="Save provided credentials for future use")
    args = parser.parse_args()

    # Load/merge config
    cfg = load_config()
    if args.spotify_id:
        cfg["spotify_client_id"] = args.spotify_id
    if args.spotify_secret:
        cfg["spotify_client_secret"] = args.spotify_secret
    if args.save_credentials:
        save_config(cfg)
        print("Credentials saved to ~/.isrc_fetcher/config.json")

    if not cfg.get("spotify_client_id"):
        print("Warning: No Spotify credentials configured. Using MusicBrainz only (slower).")
        print("Run with --spotify-id and --spotify-secret, or configure via Excel add-in.")

    fetcher = ISRCFetcher(
        spotify_client_id=cfg.get("spotify_client_id"),
        spotify_client_secret=cfg.get("spotify_client_secret"),
    )

    # Open workbook
    wb = openpyxl.load_workbook(args.file)
    ws = wb.active

    # Write verification headers if not present
    if ws.cell(row=1, column=COL_STATUS).value != "ISRC_STATUS":
        ws.cell(row=1, column=COL_STATUS).value         = "ISRC_STATUS"
        ws.cell(row=1, column=COL_FOUND_TITLE).value    = "ISRC_FOUND_TITLE"
        ws.cell(row=1, column=COL_FOUND_ARTIST).value   = "ISRC_FOUND_ARTIST"
        ws.cell(row=1, column=COL_FOUND_DURATION).value = "ISRC_FOUND_DURATION"

    # Determine rows to process
    max_row = ws.max_row
    if args.rows:
        if "-" in args.rows:
            start, end = args.rows.split("-", 1)
            rows = list(range(int(start), int(end) + 1))
        else:
            rows = [int(args.rows)]
    elif args.all:
        rows = list(range(2, max_row + 1))
    else:
        print("Error: Specify --all or --rows. Use --help for details.")
        sys.exit(1)

    total = len(rows)
    found = 0
    not_found = 0
    warnings = 0
    skipped = 0

    for i, row in enumerate(rows, 1):
        title = ws.cell(row=row, column=COL_C).value
        artist = ws.cell(row=row, column=COL_D).value
        duration_raw = ws.cell(row=row, column=COL_B).value
        existing_isrc = ws.cell(row=row, column=COL_H).value

        if not title or not artist:
            continue

        # Skip rows that already have an ISRC
        if existing_isrc and str(existing_isrc).strip():
            skipped += 1
            print(f"[{i}/{total}] {artist} - {title}... SKIPPED (has ISRC)")
            continue

        title = str(title).strip()
        artist = str(artist).strip()
        duration_secs = parse_duration(duration_raw)

        print(f"[{i}/{total}] {artist} - {title}...", end=" ", flush=True)

        result = fetcher.fetch(title, artist, duration_secs)

        if result["isrc"]:
            ws.cell(row=row, column=COL_H).value = result["isrc"]
            found += 1
            if result["warning"]:
                warnings += 1
                ws.cell(row=row, column=COL_STATUS).value = "Duration mismatch"
                print(f'{result["isrc"]} ⚠ {result["warning"]}')
            else:
                ws.cell(row=row, column=COL_STATUS).value = "Exact match"
                print(result["isrc"])

            matched = result.get("matched")
            if matched:
                ws.cell(row=row, column=COL_FOUND_TITLE).value  = matched.get("name", "")
                ws.cell(row=row, column=COL_FOUND_ARTIST).value = matched.get("artist", "")
                dur_ms = matched.get("duration_ms")
                if dur_ms:
                    ws.cell(row=row, column=COL_FOUND_DURATION).value = (
                        f"{dur_ms // 1000 // 60}:{dur_ms // 1000 % 60:02d}"
                    )
        else:
            not_found += 1
            ws.cell(row=row, column=COL_STATUS).value = "Not found"
            print("NOT FOUND")

    # Save
    wb.save(args.file)
    print(f"\nDone! Found: {found}, Not found: {not_found}, Warnings: {warnings}, Skipped: {skipped}")
    print(f"Results saved to {args.file}")


if __name__ == "__main__":
    main()
