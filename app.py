"""ISRC Fetcher — local HTTP server + browser UI.

Run: python3 app.py
Opens http://localhost:8765 in your browser automatically.
Works on Mac and Windows — no extra dependencies beyond openpyxl + requests.
"""
from __future__ import annotations

import collections
import json
import os
import re
import subprocess
import sys
import threading
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from isrc_fetcher.columns import get_cols, DEFAULT_COLUMNS
from isrc_fetcher.config import load_config, save_config
from isrc_fetcher.fetcher import ISRCFetcher
from isrc_fetcher.resolver import TrackResolver
from isrc_fetcher.ui import HTML_PAGE
from isrc_fetcher import cancel as cancel_module

import openpyxl

# ---------------------------------------------------------------------------
# Job state (shared between HTTP handler and background worker thread)
# ---------------------------------------------------------------------------

_lock = threading.Lock()
_log_buffer: collections.deque = collections.deque(maxlen=5000)

job_state: dict = {
    "running": False,
    "cancel": False,
    "progress": 0,
    "total": 0,
    "done": False,
    "found": 0,
    "not_found": 0,
    "warnings": 0,
    "skipped": 0,
    "current": "",
    "ai_cost": 0.0,
}

_worker_thread: threading.Thread | None = None


def _log(msg: str) -> None:
    if msg:
        print(msg, flush=True)
    # Lines prefixed [AI:raw] are console-only (verbose API responses)
    if msg and msg.startswith("[AI:raw]"):
        return
    with _lock:
        _log_buffer.append(msg)


def _log_console(msg: str) -> None:
    """Log to terminal only — not shown in UI."""
    if msg:
        print(msg, flush=True)


def _get_status() -> dict:
    with _lock:
        return {
            **{k: job_state[k] for k in (
                "running", "progress", "total", "done",
                "found", "not_found", "warnings", "skipped", "current", "cancel", "ai_cost",
            )},
            "log": list(_log_buffer),
        }


def _parse_duration(raw) -> int | None:
    if raw is None:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})$", str(raw).strip())
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


# ---------------------------------------------------------------------------
# Background worker
# ---------------------------------------------------------------------------

def run_fetch(file_path: str, row_start: int, row_end: int) -> None:
    cancel_module.reset()
    with _lock:
        job_state.update(
            running=True, cancel=False, progress=0, done=False,
            found=0, not_found=0, warnings=0, skipped=0, current="",
        )
        _log_buffer.clear()

    try:
        cfg = load_config()
        cols = get_cols(cfg)
        source = cfg.get("source", "deezer")
        # Support both old single-account and new multi-account config
        accounts = cfg.get("spotify_accounts", [])
        if not accounts and cfg.get("spotify_client_id"):
            accounts = [{"client_id": cfg["spotify_client_id"],
                         "client_secret": cfg.get("spotify_client_secret", "")}]
        fetcher = ISRCFetcher(
            spotify_accounts=accounts,
            source=source,
            log=_log,
            verbose_log=_log_console,
        )
        _log(f"Primary source: {source.capitalize()}")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Write column headers on first run
        if ws.cell(row=1, column=cols["status"]).value != "ISRC_STATUS":
            ws.cell(row=1, column=cols["status"]).value         = "ISRC_STATUS"
            ws.cell(row=1, column=cols["found_title"]).value    = "ISRC_FOUND_TITLE"
            ws.cell(row=1, column=cols["found_artist"]).value   = "ISRC_FOUND_ARTIST"
            ws.cell(row=1, column=cols["found_duration"]).value = "ISRC_FOUND_DURATION"

        actual_end = ws.max_row if row_end == -1 else min(row_end, ws.max_row)
        rows = list(range(row_start, actual_end + 1))
        job_state["total"] = len(rows)
        _log(f"Processing {len(rows)} rows from {os.path.basename(file_path)}...")

        found = not_found = warnings = skipped = 0

        for i, row in enumerate(rows, 1):
            if job_state["cancel"]:
                _log("Cancelled by user.")
                _safe_save(wb, file_path)
                _log("Partial results saved.")
                break

            title          = ws.cell(row=row, column=cols["titulo"]).value
            artist         = ws.cell(row=row, column=cols["interpretes"]).value
            duration_raw   = ws.cell(row=row, column=cols["duracion"]).value
            existing_isrc  = ws.cell(row=row, column=cols["isrc"]).value
            job_state["progress"] = i

            if not title or not artist:
                continue

            if existing_isrc and str(existing_isrc).strip():
                skipped += 1
                job_state["skipped"] = skipped
                _log(f"[{i}/{len(rows)}] [R{row}] {artist} - {title} -> SKIPPED (has ISRC: {existing_isrc})")
                continue

            title  = str(title).strip()
            artist = str(artist).strip()
            job_state["current"] = f"{artist} — {title}"
            result = fetcher.fetch(title, artist, _parse_duration(duration_raw))

            if result["isrc"]:
                src_tag = f"[{result.get('source', '?')}]"
                ws.cell(row=row, column=cols["isrc"]).value = result["isrc"]
                found += 1
                job_state["found"] = found

                if result["warning"]:
                    warnings += 1
                    job_state["warnings"] = warnings
                    ws.cell(row=row, column=cols["status"]).value = "Duration mismatch"
                    _log(f"[{i}/{len(rows)}] [R{row}] {src_tag} {artist} - {title} -> {result['isrc']} (duration mismatch)")
                else:
                    ws.cell(row=row, column=cols["status"]).value = "Exact match"
                    _log(f"[{i}/{len(rows)}] [R{row}] {src_tag} {artist} - {title} -> {result['isrc']}")

                matched = result.get("matched")
                if matched:
                    ws.cell(row=row, column=cols["found_title"]).value  = matched.get("name", "")
                    ws.cell(row=row, column=cols["found_artist"]).value = matched.get("artist", "")
                    dur_ms = matched.get("duration_ms")
                    if dur_ms:
                        ws.cell(row=row, column=cols["found_duration"]).value = (
                            f"{dur_ms // 1000 // 60}:{dur_ms // 1000 % 60:02d}"
                        )
            else:
                not_found += 1
                job_state["not_found"] = not_found
                ws.cell(row=row, column=cols["status"]).value = "Not found"
                _log(f"[{i}/{len(rows)}] [R{row}] {artist} - {title} -> NOT FOUND")

            # Save every 5 rows to avoid losing progress
            if i % 5 == 0:
                _safe_save(wb, file_path)

        else:
            _safe_save(wb, file_path)

        _log("")
        _log(f"Done! Found: {found} | Not found: {not_found} | Duration warnings: {warnings} | Skipped: {skipped}")
        _log(f"Results saved to {os.path.basename(file_path)}")

    except cancel_module.CancelledError:
        _log("Stopped by user.")
    except Exception as e:
        _log(f"ERROR: {e}")
    finally:
        job_state["running"] = False
        job_state["done"] = True


def _safe_save(wb, file_path: str) -> None:
    """Save workbook with error handling to prevent silent data loss."""
    try:
        wb.save(file_path)
        _log(f"Excel saved: {os.path.basename(file_path)}")
    except Exception as e:
        _log(f"WARNING: Failed to save file: {e}")


# ---------------------------------------------------------------------------
# AI helpers
# ---------------------------------------------------------------------------

CHUNK_SIZE = 20  # tracks per full cycle: queries → search → eval → save


def _get_ai_key(cfg: dict) -> str:
    return cfg.get("openai_api_key", "")


# ---------------------------------------------------------------------------
# AI Resolve worker
# ---------------------------------------------------------------------------

def run_resolve(file_path: str, row_start: int = 2, row_end: int = -1) -> None:
    """Second pass: AI-resolve tracks with 'Not found' or 'Duration mismatch'."""
    cancel_module.reset()
    with _lock:
        job_state.update(
            running=True, cancel=False, progress=0, done=False,
            found=0, not_found=0, warnings=0, skipped=0, current="AI Resolve", ai_cost=0.0,
        )
        _log_buffer.clear()

    try:
        cfg = load_config()
        ai_key = _get_ai_key(cfg)
        if not ai_key:
            _log("ERROR: No OpenAI API key configured")
            return

        accounts = cfg.get("spotify_accounts", [])
        if not accounts and cfg.get("spotify_client_id"):
            accounts = [{"client_id": cfg["spotify_client_id"],
                         "client_secret": cfg.get("spotify_client_secret", "")}]

        resolver = TrackResolver(
            ai_api_key=ai_key,
            spotify_accounts=accounts,
            log=_log,
        )

        cols = get_cols(cfg)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Collect rows that need resolution (respect row range filter)
        actual_end = ws.max_row if row_end == -1 else min(row_end, ws.max_row)
        tracks_to_resolve = []
        for row in range(row_start, actual_end + 1):
            status = ws.cell(row=row, column=cols["status"]).value
            if status in ("Not found", "Duration mismatch"):
                title = ws.cell(row=row, column=cols["titulo"]).value
                artist = ws.cell(row=row, column=cols["interpretes"]).value
                duration_raw = ws.cell(row=row, column=cols["duracion"]).value
                if title and artist:
                    dur = None
                    if duration_raw:
                        m = re.match(r"^(\d{1,2}):(\d{2})$", str(duration_raw).strip())
                        if m:
                            dur = f"{m.group(1)}:{m.group(2)}"
                    tracks_to_resolve.append({
                        "row": row,
                        "title": str(title).strip(),
                        "artist": str(artist).strip(),
                        "duration": dur,
                        "status": status,
                    })

        if not tracks_to_resolve:
            _log("No tracks to resolve — all tracks already have ISRCs or exact matches.")
            return

        job_state["total"] = len(tracks_to_resolve)
        _log(f"Found {len(tracks_to_resolve)} tracks to resolve in {os.path.basename(file_path)}")

        from isrc_fetcher.ai_client import create_ai_client
        ai = create_ai_client(ai_key, log=_log)

        # Write column headers for cleaned columns once
        if ws.cell(row=1, column=cols["cleaned_title"]).value != "AI_CLEANED_TITLE":
            ws.cell(row=1, column=cols["cleaned_title"]).value = "AI_CLEANED_TITLE"
            ws.cell(row=1, column=cols["cleaned_artist"]).value = "AI_CLEANED_ARTIST"

        total_resolved = 0
        total_chunks = (len(tracks_to_resolve) + CHUNK_SIZE - 1) // CHUNK_SIZE

        for chunk_idx in range(0, len(tracks_to_resolve), CHUNK_SIZE):
            if job_state["cancel"]:
                break

            chunk = tracks_to_resolve[chunk_idx:chunk_idx + CHUNK_SIZE]
            chunk_num = chunk_idx // CHUNK_SIZE + 1
            _log(f"[AI Resolve] ── Chunk {chunk_num}/{total_chunks} ({len(chunk)} tracks) ──")

            # Phase A: AI generates search queries for this chunk
            job_state["current"] = f"Chunk {chunk_num}/{total_chunks}: generating queries..."
            batch_input = [{"index": t["row"], "title": t["title"], "artist": t["artist"]} for t in chunk]
            query_results = ai.clean_batch(batch_input)
            query_map = {r["index"]: r for r in query_results}

            for t in chunk:
                qr = query_map.get(t["row"], {})
                if qr.get("changed"):
                    ws.cell(row=t["row"], column=cols["cleaned_title"]).value = qr["title"]
                    ws.cell(row=t["row"], column=cols["cleaned_artist"]).value = qr["artist"]
                    t["title"] = qr["title"]
                    t["artist"] = qr["artist"]
                t["queries"] = qr.get("queries", [])
                if t["queries"]:
                    _log(f"[AI] R{t['row']} queries: {' | '.join(t['queries'])}")

            job_state["ai_cost"] = ai.cost_usd

            # Phase B: API search for candidates
            job_state["current"] = f"Chunk {chunk_num}/{total_chunks}: searching APIs..."
            track_candidates = []
            for i, track in enumerate(chunk):
                if job_state["cancel"]:
                    break
                job_state["progress"] = chunk_idx + i + 1
                n = chunk_idx + i + 1

                if track.get("queries"):
                    candidates = resolver._search_with_queries(track["queries"])
                else:
                    candidates = resolver._fallback_search(track["title"], track["artist"])

                if candidates:
                    track_candidates.append({"track": track, "candidates": candidates})
                    _log(f"[AI Resolve] [{n}/{len(tracks_to_resolve)}] {track['artist']} - {track['title']} → {len(candidates)} candidates")
                else:
                    _log(f"[AI Resolve] [{n}/{len(tracks_to_resolve)}] {track['artist']} - {track['title']} → no candidates")

            if not track_candidates:
                _log(f"[AI Resolve] Chunk {chunk_num}: no candidates found, skipping eval")
                continue

            # Phase C: AI evaluation
            job_state["current"] = f"Chunk {chunk_num}/{total_chunks}: AI evaluating..."
            eval_input = [
                {
                    "index": item["track"]["row"],
                    "title": item["track"]["title"],
                    "artist": item["track"]["artist"],
                    "duration": item["track"].get("duration"),
                    "candidates": item["candidates"],
                }
                for item in track_candidates
            ]
            decisions = ai.evaluate_batch(eval_input)
            job_state["ai_cost"] = ai.cost_usd

            chunk_resolved = 0
            for decision in decisions:
                if decision["pick"] is None:
                    continue
                item = next((c for c in track_candidates if c["track"]["row"] == decision["index"]), None)
                if not item:
                    continue
                pick_idx = decision["pick"]
                if pick_idx >= len(item["candidates"]):
                    continue
                candidate = item["candidates"][pick_idx]
                row = decision["index"]
                ws.cell(row=row, column=cols["isrc"]).value = candidate["isrc"]
                ws.cell(row=row, column=cols["status"]).value = f"AI resolved ({decision['confidence']})"
                ws.cell(row=row, column=cols["found_title"]).value = candidate.get("name", "")
                ws.cell(row=row, column=cols["found_artist"]).value = candidate.get("artist", "")
                dur_ms = candidate.get("duration_ms", 0)
                if dur_ms:
                    ws.cell(row=row, column=cols["found_duration"]).value = (
                        f"{dur_ms // 1000 // 60}:{dur_ms // 1000 % 60:02d}"
                    )
                chunk_resolved += 1
                total_resolved += 1
                _log(f"[AI Resolve] R{row}: → {candidate['isrc']} ({decision['confidence']}) — {decision['reason']}")

            # Save after every chunk
            _safe_save(wb, file_path)
            job_state["found"] = total_resolved
            job_state["not_found"] = (chunk_idx + len(chunk)) - total_resolved
            _log(f"[AI Resolve] Chunk {chunk_num} saved — {chunk_resolved} resolved. Running total: {total_resolved}/{len(tracks_to_resolve)}")

        job_state["progress"] = len(tracks_to_resolve)
        _log("")
        _log(f"AI Resolve done! Resolved: {total_resolved} / {len(tracks_to_resolve)}")
        _log(f"AI cost: ${ai.cost_usd:.4f} ({ai.tokens_in} in / {ai.tokens_out} out tokens)")
        _log(f"Results saved to {os.path.basename(file_path)}")

    except cancel_module.CancelledError:
        _log("Stopped by user.")
    except Exception as e:
        _log(f"ERROR: {e}")
    finally:
        job_state["running"] = False
        job_state["done"] = True


# ---------------------------------------------------------------------------
# Native file dialog
# ---------------------------------------------------------------------------

def _open_file_dialog() -> str | None:
    if sys.platform == "darwin":
        script = (
            'set theFile to choose file of type {"xlsx", "xls", "xlsm"} '
            'with prompt "Select Excel file"\n'
            'return POSIX path of theFile'
        )
        try:
            r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True, timeout=120)
            if r.returncode == 0:
                return r.stdout.strip()
        except Exception:
            pass

    elif sys.platform == "win32":
        ps = (
            '[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; '
            'Add-Type -AssemblyName System.Windows.Forms; '
            '$d = New-Object System.Windows.Forms.OpenFileDialog; '
            '$d.Filter = "Excel files (*.xlsx;*.xls;*.xlsm)|*.xlsx;*.xls;*.xlsm"; '
            '$d.Title = "Select Excel file"; '
            'if ($d.ShowDialog() -eq "OK") { $d.FileName }'
        )
        try:
            r = subprocess.run(
                ["powershell", "-NoProfile", "-NoLogo", "-Command", ps],
                capture_output=True, timeout=120,
            )
            if r.returncode == 0 and r.stdout.strip():
                return r.stdout.decode("utf-8-sig").strip()
        except Exception:
            pass

    return None


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------

class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass  # suppress default request logging

    def _json(self, data, status=200):
        body = json.dumps(data).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _html(self, content):
        body = content.encode()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _body(self):
        n = int(self.headers.get("Content-Length", 0))
        return self.rfile.read(n) if n else b""

    def do_GET(self):
        if self.path == "/":
            self._html(HTML_PAGE)
        elif self.path == "/api/config":
            self._json(load_config())
        elif self.path == "/api/status":
            self._json(_get_status())
        else:
            self.send_error(404)

    def do_POST(self):
        global _worker_thread
        body = self._body()

        if self.path == "/api/config":
            try:
                save_config(json.loads(body))
            except (json.JSONDecodeError, ValueError) as e:
                self._json({"error": f"Invalid JSON: {e}"}, 400)
                return
            self._json({"ok": True})

        elif self.path == "/api/fetch":
            if job_state["running"]:
                self._json({"error": "Already running"}, 409)
                return
            try:
                data = json.loads(body)
            except (json.JSONDecodeError, ValueError) as e:
                self._json({"error": f"Invalid JSON: {e}"}, 400)
                return
            _worker_thread = threading.Thread(
                target=run_fetch,
                args=(data["file"], data.get("row_start", 2), data.get("row_end", -1)),
            )
            _worker_thread.start()
            self._json({"ok": True})

        elif self.path == "/api/cancel":
            job_state["cancel"] = True
            cancel_module.cancel()
            self._json({"ok": True})

        elif self.path == "/api/resolve":
            if job_state["running"]:
                self._json({"error": "Already running"}, 409)
                return
            try:
                data = json.loads(body)
            except (json.JSONDecodeError, ValueError) as e:
                self._json({"error": f"Invalid JSON: {e}"}, 400)
                return
            _worker_thread = threading.Thread(
                target=run_resolve,
                args=(data["file"], data.get("row_start", 2), data.get("row_end", -1)),
            )
            _worker_thread.start()
            self._json({"ok": True})

        elif self.path == "/api/browse":
            self._json({"path": _open_file_dialog()})

        else:
            self.send_error(404)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

class ReusableHTTPServer(HTTPServer):
    allow_reuse_address = True


def main():
    port = 8765
    server = ReusableHTTPServer(("127.0.0.1", port), Handler)
    url = f"http://localhost:{port}"
    print(f"ISRC Fetcher running at {url}")
    print("Press Ctrl+C to stop.\n")
    webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down...")
        server.server_close()
        if _worker_thread and _worker_thread.is_alive():
            print("Waiting for current job to finish (Ctrl+C again to force)...")
            job_state["cancel"] = True
            cancel_module.cancel()
            _worker_thread.join(timeout=10)
        print("Stopped.")


if __name__ == "__main__":
    main()
